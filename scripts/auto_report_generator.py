"""Auto report generator for M9: Excel parsing + console summary + optional Slack post."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib import error, request

from openpyxl import load_workbook


@dataclass
class SlackMessage:
    """Simple message payload model for Slack notifications."""

    channel: str
    text: str


def _normalize_header(value: Any, idx: int) -> str:
    if value is None:
        return f"column_{idx + 1}"
    text = str(value).strip()
    return text or f"column_{idx + 1}"


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_excel_rows(excel_path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Load rows from an Excel file and return normalized dictionaries."""
    path = Path(excel_path)
    workbook = load_workbook(filename=path, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    headers = [_normalize_header(value, idx) for idx, value in enumerate(rows[0])]
    parsed_rows: list[dict[str, Any]] = []

    for row in rows[1:]:
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_dict: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            row_dict[header] = row[idx] if idx < len(row) else None
        parsed_rows.append(row_dict)

    return parsed_rows


def build_slack_messages(rows: list[dict[str, Any]], default_channel: str) -> list[SlackMessage]:
    """Transform parsed rows into Slack summary message payloads."""
    if not rows:
        return [SlackMessage(channel=default_channel, text="AX 자동 리포트: 데이터 행이 없습니다.")]

    headers = list(rows[0].keys())
    lines = [
        "AX 자동 리포트",
        f"- 총 데이터 행: {len(rows)}",
        f"- 컬럼 수: {len(headers)}",
        f"- 컬럼 목록: {', '.join(headers)}",
    ]

    numeric_headers: list[str] = []
    for header in headers:
        numeric_count = sum(1 for row in rows if _to_float(row.get(header)) is not None)
        if numeric_count > 0:
            numeric_headers.append(header)

    for header in numeric_headers[:3]:
        values = [_to_float(row.get(header)) for row in rows]
        clean_values = [v for v in values if v is not None]
        if not clean_values:
            continue
        avg = sum(clean_values) / len(clean_values)
        lines.append(
            f"- {header}: min={min(clean_values):.2f}, avg={avg:.2f}, max={max(clean_values):.2f}"
        )

    sample = rows[0]
    sample_preview = ", ".join(f"{k}={sample.get(k)}" for k in headers[:3])
    lines.append(f"- 첫 행 샘플: {sample_preview}")

    return [SlackMessage(channel=default_channel, text="\n".join(lines))]


def post_to_slack(webhook_url: str, messages: list[SlackMessage]) -> None:
    """Send messages to Slack webhook endpoint."""
    if not webhook_url.strip():
        print("Slack 전송 건너뜀: webhook_url이 비어 있습니다.")
        return

    for message in messages:
        payload = {"text": message.text, "channel": message.channel}
        data = json.dumps(payload).encode("utf-8")
        req = request.Request(webhook_url, data=data, headers={"Content-Type": "application/json"})
        try:
            with request.urlopen(req, timeout=10) as resp:
                status = getattr(resp, "status", 200)
            print(f"Slack 전송 완료: status={status}, channel={message.channel}")
        except error.URLError as exc:
            print(f"Slack 전송 실패: {exc}")


def run_pipeline(
    excel_path: str | Path,
    webhook_url: str,
    sheet_name: str | None = None,
    default_channel: str = "#general",
) -> None:
    """High-level pipeline: parse Excel, build messages, print summary, and optionally post."""
    rows = load_excel_rows(excel_path=excel_path, sheet_name=sheet_name)
    messages = build_slack_messages(rows=rows, default_channel=default_channel)

    for message in messages:
        print("===== AX REPORT SUMMARY =====")
        print(message.text)
        print("=============================")

    post_to_slack(webhook_url=webhook_url, messages=messages)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate AX report from Excel and optionally post to Slack.")
    parser.add_argument("excel_path", help="Path to .xlsx file")
    parser.add_argument("--sheet", dest="sheet_name", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--channel", dest="default_channel", default="#general", help="Slack channel")
    parser.add_argument("--webhook-url", dest="webhook_url", default="", help="Slack incoming webhook URL")
    args = parser.parse_args()

    run_pipeline(
        excel_path=args.excel_path,
        webhook_url=args.webhook_url,
        sheet_name=args.sheet_name,
        default_channel=args.default_channel,
    )
